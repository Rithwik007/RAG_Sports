from pathlib import Path
from typing import List, Dict, Any
import pypdf
from docx import Document
import json
import csv


class DocumentLoader:
    """Base class for document loaders"""
    
    # Reads a PDF file page-by-page and extracts text content from each non-empty page.
    @staticmethod
    def load_pdf(file_path: Path) -> List[Dict[str, Any]]:
        """Load PDF document"""
        pages = []
        with open(file_path, 'rb') as file:
            pdf_reader = pypdf.PdfReader(file)
            for page_num, page in enumerate(pdf_reader.pages):
                text = page.extract_text()
                if text.strip():
                    pages.append({
                        "page_number": page_num + 1,
                        "content": text,
                        "source": str(file_path)
                    })
        return pages
    
    # Reads a Word (.docx) file, concatenates all paragraphs, and returns it as a single-page document.
    @staticmethod
    def load_docx(file_path: Path) -> List[Dict[str, Any]]:
        """Load DOCX document"""
        pages = []
        doc = Document(file_path)
        full_text = []
        
        for paragraph in doc.paragraphs:
            if paragraph.text.strip():
                full_text.append(paragraph.text)
        
        # Treat entire document as one page for simplicity
        if full_text:
            pages.append({
                "page_number": 1,
                "content": "\n".join(full_text),
                "source": str(file_path)
            })
        
        return pages
    
    # Reads a Markdown (.md) file and returns its full text content as a single-page document.
    @staticmethod
    def load_markdown(file_path: Path) -> List[Dict[str, Any]]:
        """Load Markdown document"""
        with open(file_path, 'r', encoding='utf-8') as file:
            content = file.read()
        
        return [{
            "page_number": 1,
            "content": content,
            "source": str(file_path)
        }]
    
    # Reads a CSV file row-by-row using DictReader and joins all rows into a single text document.
    @staticmethod
    def load_csv(file_path: Path) -> List[Dict[str, Any]]:
        """Load CSV document"""
        content = []
        with open(file_path, 'r', encoding='utf-8') as file:
            csv_reader = csv.DictReader(file)
            for row in csv_reader:
                content.append(str(row))
        
        full_content = "\n".join(content)
        
        return [{
            "page_number": 1,
            "content": full_content,
            "source": str(file_path)
        }]
    
    # Reads a JSON file and pretty-prints it as indented text content.
    @staticmethod
    def load_json(file_path: Path) -> List[Dict[str, Any]]:
        """Load JSON document"""
        with open(file_path, 'r', encoding='utf-8') as file:
            data = json.load(file)
        
        content = json.dumps(data, indent=2)
        
        return [{
            "page_number": 1,
            "content": content,
            "source": str(file_path)
        }]

    # Reads an Excel (.xlsx, .xls) document sheet by sheet using openpyxl.
    @staticmethod
    def load_excel(file_path: Path) -> List[Dict[str, Any]]:
        """Load Excel (.xlsx, .xls) document"""
        import openpyxl
        
        pages = []
        wb = openpyxl.load_workbook(file_path, data_only=True)
        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            sheet = wb[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            
            headers = [str(cell) if cell is not None else "" for cell in rows[0]]
            sheet_lines = [f"Sheet: {sheet_name}"]
            
            for row in rows[1:]:
                if not any(row):
                    continue
                row_dict = {}
                for h, cell in zip(headers, row):
                    if cell is not None and str(cell).strip():
                        row_dict[h] = cell
                if row_dict:
                    sheet_lines.append(str(row_dict))
            
            if len(sheet_lines) > 1:
                pages.append({
                    "page_number": sheet_idx + 1,
                    "content": "\n".join(sheet_lines),
                    "source": str(file_path)
                })
        
        return pages
    
    # Routes a file to the correct loader based on its extension (.pdf, .docx, .md, .csv, .json, .xlsx, .xls).
    @classmethod
    def load_document(cls, file_path: Path) -> List[Dict[str, Any]]:
        """Load document based on file extension"""
        suffix = file_path.suffix.lower()
        
        loaders = {
            '.pdf': cls.load_pdf,
            '.docx': cls.load_docx,
            '.md': cls.load_markdown,
            '.csv': cls.load_csv,
            '.json': cls.load_json,
            '.xlsx': cls.load_excel,
            '.xls': cls.load_excel,
        }
        
        loader = loaders.get(suffix)
        if loader is None:
            raise ValueError(f"Unsupported file type: {suffix}")
        
        return loader(file_path)
